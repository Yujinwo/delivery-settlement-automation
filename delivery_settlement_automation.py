import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ==============================
# 0. 기본 설정
# ==============================
INPUT_DIR = "input"
OUTPUT_DIR = "output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "settlement_report.xlsx")

# ==============================
# 1. 입력 폴더 확인
# ==============================
os.makedirs(OUTPUT_DIR, exist_ok=True)

files = glob.glob(os.path.join(INPUT_DIR, "*.csv"))

if not files:
    raise FileNotFoundError(
        "input 폴더에 CSV 파일이 없습니다. 주문 데이터를 넣어주세요."
    )

print(f"[INFO] 발견된 파일 수: {len(files)}")


# ==============================
# 2. CSV 취합
# ==============================
df_list = []

for file in files:
    try:
        temp = pd.read_csv(file)
        df_list.append(temp)
        print(f"[INFO] 로드 완료: {file}")
    except Exception as e:
        print(f"[WARN] 파일 로드 실패: {file} / {e}")

df = pd.concat(df_list, ignore_index=True)
#df = df.drop_duplicates(subset=["order_id"])

print(f"[INFO] 총 데이터 건수: {len(df)}")


# ==============================
# 3. 컬럼 검증
# ==============================
required_cols = {
    "order_id",
    "order_date",
    "store_name",
    "menu_name",
    "qty",
    "unit_price",
    "fee_rate",
}

missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"필수 컬럼 누락: {missing}")


# ==============================
# 데이터 타입 정제 (추가)
# ==============================
df["order_date"] = pd.to_datetime(df["order_date"], errors="coerce")

numeric_cols = ["qty", "unit_price", "fee_rate"]
for col in numeric_cols:
    df[col] = pd.to_numeric(df[col], errors="coerce")

# 결측 제거
before = len(df)
df = df.dropna(subset=["order_id", "order_date", "qty", "unit_price", "fee_rate"])
after = len(df)

print(f"[INFO] 결측 제거: {before - after}건")


# ==============================
# 데이터 이상치 검증 (추가)
# ==============================
if (df["qty"] <= 0).any():
    raise ValueError("수량 오류 데이터 존재")

if (df["unit_price"] <= 0).any():
    raise ValueError("단가 오류 데이터 존재")

if not df["fee_rate"].between(0, 1).all():
    raise ValueError("수수료율 범위 오류")

# 중복 주문 처리
before = len(df)
df = df.drop_duplicates(subset=["order_id"])
after = len(df)
print(f"[INFO] 중복 제거: {before - after}건")

# ==============================
# 4. 정산 계산
# ==============================
df["gross_amount"] = df["qty"] * df["unit_price"]
df["fee_amount"] = (df["gross_amount"] * df["fee_rate"]).round(0)
df["settlement_amount"] = df["gross_amount"] - df["fee_amount"]

print("[INFO] 정산 계산 완료")


# ==============================
# 5. 가맹점별 집계
# ==============================
store_summary = (
    df.groupby("store_name")[["gross_amount", "fee_amount", "settlement_amount"]]
    .sum()
    .reset_index()
)

# ==============================
# 6. 날짜별 매출
# ==============================
date_summary = (
    df.groupby("order_date")["gross_amount"].sum().reset_index()
)


# ==============================
# 7. 엑셀 저장
# ==============================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="raw_data", index=False)
    store_summary.to_excel(writer, sheet_name="by_store", index=False)
    date_summary.to_excel(writer, sheet_name="by_date", index=False)

print(f"[INFO] 엑셀 생성 완료: {OUTPUT_FILE}")


# ==============================
# 8. openpyxl 서식 적용
# ==============================
wb = load_workbook(OUTPUT_FILE)

header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
center_align = Alignment(horizontal="center")


def style_sheet(ws, money_cols=None):
    # 헤더 스타일
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 열 너비 자동 조정
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # 금액 컬럼 서식
    if money_cols:
        headers = [cell.value for cell in ws[1]]
        for col_name in money_cols:
            if col_name in headers:
                idx = headers.index(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    for cell in row:
                        cell.number_format = "#,##0"


# 시트별 스타일 적용
style_sheet(
    wb["raw_data"],
    money_cols=["gross_amount", "fee_amount", "settlement_amount"],
)

style_sheet(
    wb["by_store"],
    money_cols=["gross_amount", "fee_amount", "settlement_amount"],
)

style_sheet(
    wb["by_date"],
    money_cols=["gross_amount"],
)

wb.save(OUTPUT_FILE)

print("[INFO] 엑셀 서식 적용 완료")
print("[DONE] 정산 자동화 작업 완료")